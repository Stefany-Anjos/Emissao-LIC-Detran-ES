from dotenv import load_dotenv
from openpyxl import load_workbook
from twocaptcha import TwoCaptcha
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import pyautogui
import requests
import shutil
import psutil
import pdfplumber
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import undetected_chromedriver as uc
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import os
import re
import pickle
import base64
from PIL import Image

# Cria uma pasta de saída com base no nome da planilha
pasta_baixou = r"C:\Users\stefany\Downloads"

# Apy para quebra de captcha
API_KEY = os.getenv("api_key")

# Cria uma pasta de saída com base no nome da planilha


def criar_pasta_saida(caminho_planilha, pasta_downloads):
    nome_arquivo = os.path.splitext(os.path.basename(caminho_planilha))[0]
    pasta_saida = os.path.join(pasta_downloads, nome_arquivo)

    # Verifica se a pasta existe
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
        print(f"Pasta criada: {pasta_saida}")
    else:
        print(f"A pasta já existe: {pasta_saida}")

    return pasta_saida


def caminho_paraBoleto(pasta_baixou, pasta_saida, placa_atual):
    # Verifica o nome de todos os processos abertos
    for process in psutil.process_iter(['pid', 'name']):
        # Acha o Acrobat na lista de processos
        if 'Acrobat' in process.info['name']:
            process.terminate()  # Fecha o Adobe
    # Verifica os processos ate encontrar arquivos com a extensão PDF
    arquivos = [f for f in os.listdir(pasta_baixou) if f.endswith('.pdf')]
    # Obtem o ultimo arquivo baixado
    arquivo = max(arquivos, key=lambda f: os.path.getmtime(
        os.path.join(pasta_baixou, f)))
    # Verifica a lista de processos
    if arquivo in arquivos:
        # Define o caminho para salvar o arquivo
        caminho_arquivo = os.path.join(pasta_baixou, arquivo)
        # Renomeia o arquivo com a placa
        novo_nome = os.path.join(pasta_saida, f"{placa_atual}.pdf")
        # Move o arquivo para a pasta correta
        shutil.move(caminho_arquivo, novo_nome)
        print(f"Boleto salvo como {novo_nome}")
        # Aguarda 2 segundos
        time.sleep(2)
        # Pegar todas as janelas abertas
        abas = navegador.window_handles
        # Fecha a página aberta
        pyautogui.hotkey('ctrl', 'w')
        # Aguarda 2 segundos
        time.sleep(2)
        # Voltar para a aba original
        navegador.switch_to.window(abas[0])
        # Aguarda 2 segundos
        time.sleep(2)
        # Atualiza a página para uma nova pesquisa
        navegador.refresh()
        # Aguarda 2 segundos
        time.sleep(2)
        return novo_nome


# Enviar requisição para resolver o captcha do IPVA
def enviar_requisicao_captcha_1(api_key, base64_image):
    url = "http://2captcha.com/in.php"
    data = {
        "key": api_key,
        "method": "base64",
        "body": base64_image,
        "json": 1,
    }

    # Envia a resposta do captcha
    response = requests.post(url, data=data).json()
    if response["status"] == 1:  # Status 1 indica sucesso
        captcha_id = response["request"]
        return captcha_id
    else:
        raise Exception(f"Erro ao enviar captcha: {response}")

# Obter o token de resposta do CAPTCHA


def obter_resposta_captcha_1(api_key, captcha_id):
    url = f"http://2captcha.com/res.php?key={
        api_key}&action=get&id={captcha_id}&json=1"

    while True:
        response = requests.get(url).json()

        if response["status"] == 1:
            return response["request"]
        elif response["request"] == "CAPCHA_NOT_READY":
            print("Captcha ainda não resolvido, aguardando...")
        else:
            raise Exception(f"Erro ao obter resposta: {response}")

        time.sleep(5)


# Caminho da pasta de saida de arquivos
pasta_downloads = r"C:\Users\stefany\Desktop\Detran ES"
# Conexão com os itens ocultos
load_dotenv()
# Definir o caminho da planilha
caminho_planilha = r"C:\Users\stefany\Desktop\Detran ES\Teste LIC ES.xlsx"

# Cria a pasta de saida como nome da planilha
pasta_saida = criar_pasta_saida(caminho_planilha, pasta_downloads)

# Abre a planilha do Excel
planilha = load_workbook(caminho_planilha)

# Passa a instância da planilha BASE
guia_dados = planilha['BASE']

# Passa os cabeçalhos
guia_dados.cell(row=1, column=1, value="STATUS")
guia_dados.cell(row=1, column=2, value="PLACA")
guia_dados.cell(row=1, column=3, value="RENAVAM")
guia_dados.cell(row=1, column=4, value="VALOR")
guia_dados.cell(row=1, column=5, value="VENCIMENTO")

index = 0
linhas = list(guia_dados.iter_rows(min_row=2, max_row=guia_dados.max_row))

service_obj = Service(ChromeDriverManager().install())

chrome_options = Options()
chrome_options.add_argument("--disable-notifications")
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_argument(
    "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

navegador = uc.Chrome(service=service_obj, options=chrome_options)

# Entra no site do Detran ES
navegador.get(
    "https://publicodetran.es.gov.br/boletolicenciamento/consultaBoletoLicenciamento.asp")

# Abre em tela cheia
navegador.maximize_window()

try:
    cont = 1  # Inicializa a contagem de arquivos a serem salvos

    # Loop para percorrer todas as linhas da planilha
    while index < len(linhas):

        erro_captcha = False

        row = linhas[index]  # Quantidades de linhas da planilha
        status_atual = row[0].value  # Obtem o status
        placa_atual = row[1].value  # Obtem a placa
        renavam_atual = row[2].value  # Obtem o renavam

        # Verificar se o status está vazio ou None
        if not status_atual:
            # Preenche o campo placa com a placa obtida da planilha
            campo_placa = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.ID, "placa")))
            # Limpa o campo placa
            campo_placa.clear()
            campo_placa.send_keys(placa_atual)

            # Aguarda 2 segundos
            time.sleep(2)
            
            # Preenche o campo renavam com o renavam obtida da planilha
            campo_renavam = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.ID, "renavam")))
            # Limpa o campo renavam
            campo_renavam.clear()
            campo_renavam.send_keys(renavam_atual)

            # Quebrar captcha
            while erro_captcha is False:
                # Esperar e localizar a imagem do CAPTCHA
                imagem_captcha = WebDriverWait(navegador, 10).until(
                    EC.visibility_of_element_located((By.ID, "imgCaptcha"))
                )

                # Fazer o screenshot de toda a página
                navegador.save_screenshot("screenshot.png")

                # Obter as dimensões do elemento do CAPTCHA
                location = imagem_captcha.location  # Posição do elemento
                # Tamanho do elemento (largura e altura)
                size = imagem_captcha.size

                # Abrir o screenshot e recortar o CAPTCHA
                with Image.open("screenshot.png") as img:
                    left = location['x']
                    top = location['y']
                    right = left + size['width']
                    bottom = top + size['height']
                    captcha = img.crop((left, top, right, bottom))

                    # Salvar o recorte do CAPTCHA
                    captcha.save("captcha_cropped.png")

                # Converter a imagem recortada para base64
                with open("captcha_cropped.png", "rb") as image_file:
                    base64_image = base64.b64encode(
                        image_file.read()).decode("utf-8")

                # Enviar CAPTCHA para o 2Captcha
                captcha_id = enviar_requisicao_captcha_1(API_KEY, base64_image)

                # Obter a resposta do CAPTCHA
                resposta_captcha = obter_resposta_captcha_1(
                    API_KEY, captcha_id)
                print(f"Resposta do CAPTCHA: {resposta_captcha}")

                # Inserir a resposta no campo do CAPTCHA
                campo_captcha = WebDriverWait(navegador, 10).until(
                    EC.visibility_of_element_located((By.ID, "txtCaptcha"))
                )

                # Limpa o campo resposta capctha
                campo_captcha.clear()

                # Aguarda 1 segundo
                time.sleep(1)

                # Digitar a resposta caractere por caractere
                for char in resposta_captcha:
                    campo_captcha.send_keys(char)
                    time.sleep(0.2)

                # Clique no botão OK
                botao_ok = navegador.find_element(By.ID, "btnSubmit")
                botao_ok.click()

                # Aguarda 2 segundos
                time.sleep(2)

                try:
                    # Tenta localizar as telas de IPVA
                    tela_ipva = WebDriverWait(navegador, 10).until(
                        EC.visibility_of_element_located(
                            (By.CSS_SELECTOR, "#LicenciamentoExercicio > table > tbody"))
                    )
                    erro_captcha = True
                    break
                except TimeoutException:
                    try:
                        # Tenta localizar a tela nada consta
                        tela_ND = WebDriverWait(navegador, 10).until(
                            EC.visibility_of_element_located(
                                (By.CSS_SELECTOR, "body > table"))
                        )
                        erro_captcha = True
                        break
                    except TimeoutException:
                         try:
                            # Verifica se a página tem renavam não pertence a placa
                            tela_erroRenavam = WebDriverWait(navegador, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "body > center > font")))
                            tela_erroRenavam = tela_erroRenavam.text
                            erro_captcha = True
                            break
                         
                         # Não achou nenhum das telas
                         except TimeoutException:
                            erro_captcha = False
            # Aguarda 2 segundos
            time.sleep(2)

            try:
                # Espera até que o elemento esteja visível e disponível para interação
                campo_ND = WebDriverWait(navegador, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "body > table")))

                # Obtem a informação de nada consta
                valor_ND = campo_ND.text

                # Verifica se tem débitos para a placa
                if campo_ND == ("Nenhum débito do tipo escolhido em aberto cadastrado para este veículo."):
                    print(f"Sem débito para a placa {
                          placa_atual}, pulando para a próxima.")
                guia_dados.cell(row=index + 2, column=1, value="Sem débito")

                navegador.refresh()

                # Salvar após atualizar o status na planilha
                planilha.save(caminho_planilha)
                continue  # Pula para a próxima iteração do loop

            except TimeoutException:
                campo_lic = ""  # Inicializa a variável campo_lic
                id_linha = 2  # variável que será incrementada no while
                botao_boleto = ""
                id_linha2 = 19
                while campo_lic != "Licenciamento Anual 2025":
                    campo_lic = navegador.find_element(
                        By.CSS_SELECTOR, f"#LicenciamentoExercicio > table > tbody > tr:nth-child({id_linha}) > td:nth-child(1)")
                    campo_lic = campo_lic.text
                    id_linha += 2
                    if campo_lic == "  Licenciamento Anual 2025":
                        break
                while True:
                    try:
                        botao_boleto = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable(
                            (By.CSS_SELECTOR, f"#LicenciamentoExercicio > table > tbody > tr:nth-child({id_linha2}) > td > input[type=button]")))
                        # Verifica se o texto do botão é "Emitir Boleto"
                        if botao_boleto.get_attribute("value") == "Emitir Boleto":
                           # Clique no botão emitir
                           botao_emitir = navegador.find_element(
                           By.CSS_SELECTOR, f"#LicenciamentoExercicio > table > tbody > tr:nth-child({id_linha2}) > td > input[type=button]")
                           botao_emitir.click()
                           # Alternar para a última aba aberta
                           navegador.switch_to.window(navegador.window_handles[-1])
                           tim.sleep()
                           erro_404 = WebDriverWait(navegador, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > div > fieldset > h2")))
                           erro_404 = erro_404.text
                           if erro_404 == "404 - File or directory not found.":
                              guia_dados.cell(row=index + 2, column=1, value="Erro ao emitir")
                           break  # Sai do loop se encontrar o botão correto
                    except TimeoutException:
                        id_linha2 += 1
                    continue
            except TimeoutException:
                guia_dados.cell(row=index + 2, column=1, value=f"{tela_erroRenavam}")
                break  
            try:
                # Pegar todas as abas ou janelas abertas
                abas = navegador.window_handles
                # Alternar para a última aba (ou janela) aberta
                navegador.switch_to.window(abas[-1])
                # Esperar tela de atenção para clicar no botão entrar com o GOV
                botaoAcessarGOV = WebDriverWait(navegador, 60).until(
                        EC.element_to_be_clickable((By.ID, "GovBr"))).click()

                # Aguarda 5 segundos
                time.sleep(5)

                # Verifica se o campo de CPF está disponivel
                campoCPF = WebDriverWait(navegador, 60).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#accountId"))
                    )

                # Insere o CPF no campo
                campoCPF.send_keys(os.getenv("LOGIN"))

                # Botão continuar
                botaoContinuarLogin = navegador.find_element(By.CSS_SELECTOR, "#enter-account-id")
                botaoContinuarLogin.click()

                # Verifica se o capo senha está disponivel
                campoSenha = WebDriverWait(navegador, 1000).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#password"))
                    )

                # Insere a senha no campo
                campoSenha.send_keys(os.getenv("SENHA"))

                # Clica no botão entrar
                botaoEntrar = navegador.find_element(By.CSS_SELECTOR, "#submit-button")
                botaoEntrar.click()

                # Mover o mouse para a posição (x=1192, y=393) e clicar
                pyautogui.click(x=1192, y=393)

                # x,y = pyautogui.position()
                # print(f"As cordenadas do botão são: x = {x}, y= {y}")

                # Aguarda 2 segundos
                time.sleep(2)

                # Rolar até o final da página
                navegador.execute_script("window.scrollTo(0, document.body.scrollHeight);")

                # Aguarda 2 segundos
                time.sleep(2)

                # Clique no botão imprimir ou salvar PDF
                botao_baixar = navegador.find_element(By.ID, "btnPrint")
                botao_baixar.click()

                # Aguardar para observar o resultado
                time.sleep(2)

            except TimeoutException:
                print("Erro ao realizar login no gov.br")
                continue

            # Aguarda 2 segundos
            time.sleep(2)

            # Chama a função de renomear boleto e mover para a pasta correta
            caminho_paraBoleto(pasta_baixou, pasta_saida, placa_atual)

            # Atualiza o status na planilha para "Boleto baixado"
            guia_dados.cell(row=index + 2, column=1, value="OK!")
            # Atualiza e salva a planilha
            planilha.save(caminho_planilha)
            index += 1
        else:
            print("Placa já pesquisada")
            index += 1
finally:
    navegador.quit()