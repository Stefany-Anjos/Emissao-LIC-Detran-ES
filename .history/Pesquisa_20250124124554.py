from dotenv import load_dotenv
from openpyxl import load_workbook
from twocaptcha import TwoCaptcha
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import pyautogui
import requests
import shutil
import psutil
import pdfplumber
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import undetected_chromedriver as uc
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import os
import re
import pickle
import base64
from PIL import Image

# Cria uma pasta de saída com base no nome da planilha
pasta_baixou = r"C:\Users\stefany\Downloads"

API_KEY = "api_key"

# Cria uma pasta de saída com base no nome da planilha
def criar_pasta_saida(caminho_planilha, pasta_downloads):
    nome_arquivo = os.path.splitext(os.path.basename(caminho_planilha))[0]
    pasta_saida = os.path.join(pasta_downloads, nome_arquivo)

    # Verifica se a pasta existe
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
        print(f"Pasta criada: {pasta_saida}")
    else:
        print(f"A pasta já existe: {pasta_saida}")

    return pasta_saida

def caminho_paraBoleto(pasta_baixou, pasta_saida, placa_atual):
    # Verifica o nome de todos os processos abertos
    for process in psutil.process_iter(['pid', 'name']):
        # Acha o Acrobat na lista de processos
        if 'Acrobat' in process.info['name']:
            process.terminate()  # Fecha o Adobe
    arquivos = [f for f in os.listdir(pasta_baixou) if f.endswith('.pdf')]
    arquivo = max(arquivos, key=lambda f: os.path.getmtime(os.path.join(pasta_baixou, f)))
    if arquivo in arquivos:
        caminho_arquivo = os.path.join(pasta_baixou, arquivo)
        novo_nome = os.path.join(pasta_saida, f"{placa_atual}.pdf")
        # Renomeia o arquivo com a placa
        shutil.move(caminho_arquivo, novo_nome)
        print(f"Boleto salvo como {novo_nome}")
        return novo_nome

# Enviar requisição para resolver o captcha do IPVA
def enviar_requisicao_captcha_1(api_key, base64_image):
    url = "http://2captcha.com/in.php"
    data = {
        "key": api_key,
        "method": "base64",
        "body": base64_image,
        "json": 1,
    }

    response = requests.post(url, data=data).json()
    if response["status"] == 1:  # Status 1 indica sucesso
        captcha_id = response["request"]
        return captcha_id
    else:
        raise Exception(f"Erro ao enviar captcha: {response}")

# Obter o token de resposta do CAPTCHA
def obter_resposta_captcha_1(api_key, captcha_id):
    url = f"http://2captcha.com/res.php?key={api_key}&action=get&id={captcha_id}&json=1"

    while True:
        response = requests.get(url).json()

        if response["status"] == 1:
            return response["request"]
        elif response["request"] == "CAPCHA_NOT_READY":
            print("Captcha ainda não resolvido, aguardando...")
        else:
            raise Exception(f"Erro ao obter resposta: {response}")

        time.sleep(5)
        
# Definir o caminho da planilha
pasta_downloads = r"C:\Users\stefany\Desktop\Detran ES"
load_dotenv()
caminho_planilha = r"C:\Users\stefany\Desktop\Detran ES\Teste LIC ES.xlsx"

pasta_saida = criar_pasta_saida(caminho_planilha, pasta_downloads)

# Abre a planilha do Excel
planilha = load_workbook(caminho_planilha)

# Passa a instância da planilha BASE
guia_dados = planilha['BASE']

# Passa os cabeçalhos
guia_dados.cell(row=1, column=1, value="STATUS")
guia_dados.cell(row=1, column=2, value="PLACA")
guia_dados.cell(row=1, column=3, value="RENAVAM")
guia_dados.cell(row=1, column=4, value="VALOR")
guia_dados.cell(row=1, column=5, value="VENCIMENTO")

index = 0
linhas = list(guia_dados.iter_rows(min_row=2, max_row=guia_dados.max_row))

service_obj = Service(ChromeDriverManager().install())

chrome_options = Options()
chrome_options.add_argument("--disable-notifications")
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_argument(
    "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")
navegador = uc.Chrome(service=service_obj, options=chrome_options)

# Entra no site do Detran ES
navegador.get("https://publicodetran.es.gov.br/boletolicenciamento/consultaBoletoLicenciamento.asp")

# Abre em tela cheia
navegador.maximize_window()

try:

finally:
    navegador


